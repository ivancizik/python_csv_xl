# https://github.com/ivancizik/python_csv_xl
# change cwd variable to change directory
# change encoding='utf-8' to change encoding for .csv files
# change quotechar='"' and delimiter=',' for quota characters and delimiter in .csv files
# change ws.title = "Data" for title of sheet in excel file

# /// import + other system stuff
import glob, os
import csv
import openpyxl

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# /// variables

cwd = os.getcwd()
cwd = str(cwd) + r"//sample files" # use this to specify other folder in script directory, comment code if files are in same folder as script
os.chdir(cwd)

list_of_files = []

csv_row_coun = 0 # progress bar counter
csv_dest_filename = "" # output file name

#CODE

print(">Searching for .csv files in", cwd)
list_of_files = []

for file in glob.glob("*.csv"):     # checks for .csv files in folder
    list_of_files.insert(0, file)   # add files to list

print("Found", len(list_of_files), ".csv files:")
print(', '.join(list_of_files))
print("")

for idx, file in enumerate(list_of_files): #opens one by one .csv files
    csv_dest_filename = str(file)
    csv_dest_filename = csv_dest_filename[:-4] + ".xlsx" # removes .csv from file and add .xlsx to output file
    print(">Converting", file, "to excel")
    f = open(file, encoding='utf-8', errors='ignore') # open csv file and specifies the encoding
    csv_row_coun = sum(1 for line in f) # counts rows in csv file, this is only for progress bar
    f.seek(0) # resets position in csv file to beginning of the file

    csv.register_dialect("colons", quotechar='"', delimiter=',') # set delimiter
    reader = csv.reader(f, dialect="colons")
    wb = Workbook()

    ws = wb.worksheets[0]
    ws.title = "Data" # title of the sheet

    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            ws['%s%s'%(column_letter, (row_index + 1))].value = cell
        #progress bar, updates progress after every 100 row
        if (row_index + 1) % 100 == 0:
            print(row_index + 1, "of ", csv_row_coun, "rows done")
        if (row_index + 1) == csv_row_coun:
            print(row_index + 1, "of ", csv_row_coun, "rows done")

    print(">Writing data to the file", csv_dest_filename, "this can take a while, please wait")
    wb.save(filename = csv_dest_filename)
    print(">Done with", file)
    print("")

print("")
